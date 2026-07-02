import csv
import json
import os
import re
from collections import Counter
from datetime import date, datetime, time, timedelta
from pathlib import Path

import openpyxl


DEFAULT_SOURCE_DIR = Path(
    r"C:\Users\Sky.Lu\Thermo Fisher Scientific\IT BA Team - Timesheet"
)
SOURCE_DIR = Path(os.environ.get("TIMESHEET_SOURCE_DIR", DEFAULT_SOURCE_DIR))
SOURCE_URL = (
    "https://thermofisher.sharepoint.com/:f:/r/sites/ITBATeam/"
    "Shared%20Documents/General/Timesheet?csf=1&web=1&e=pvWM7c"
)
OUTPUT_FILE = Path(__file__).resolve().parents[1] / "data" / "timesheet-data.js"

STANDARD_DAY_HOURS = 8
STANDARD_WEEK_HOURS = 40

DATE_FORMATS = ("%Y/%m/%d", "%m/%d/%Y", "%Y-%m-%d", "%m-%d-%Y")
TIME_FORMATS = ("%H:%M:%S", "%H:%M", "%I:%M:%S %p", "%I:%M %p")

SHOW_TIME_AS = {
    "0": "Free",
    "1": "Tentative",
    "2": "Busy",
    "3": "Out of Office",
    "4": "Working Elsewhere",
}

CATEGORY_ORDER = [
    "Project",
    "CR",
    "Mgmt",
    "Sup",
    "Other",
    "PTO",
    "Holiday",
    "Reminder",
    "Canceled",
]
WORK_CATEGORIES = {"Project", "CR", "Mgmt", "Sup", "Other"}
TIME_OFF_CATEGORIES = {"PTO", "Holiday"}
DISTRIBUTION_CATEGORIES = ["Project", "CR", "Mgmt", "Sup"]


def parse_date(value):
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    text = str(value).strip()
    for fmt in DATE_FORMATS:
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def parse_time(value):
    if value is None or value == "":
        return time(0, 0)
    if isinstance(value, datetime):
        return value.time().replace(microsecond=0)
    if isinstance(value, time):
        return value.replace(microsecond=0)

    text = str(value).strip()
    for fmt in TIME_FORMATS:
        try:
            return datetime.strptime(text, fmt).time()
        except ValueError:
            continue
    return time(0, 0)


def bool_value(value):
    return str(value or "").strip().lower() in {"true", "1", "yes"}


def normalize_subject(subject):
    return re.sub(r"\s+", " ", subject or "").strip()


def normalize_tfs_code(value):
    text = str(value or "").upper()
    match = re.search(r"\bTFS\s*[-_ ]?(\d{2,5})", text)
    if not match:
        return ""
    return f"TFS{match.group(1).zfill(3)}"


def member_from_file(path):
    stem = path.stem.lower()
    if "dai" in stem:
        return "Dai"
    if "mia" in stem:
        return "Mia"
    if "sky" in stem:
        return "Sky"
    return ""


def is_app_mapping_file(path):
    return path.suffix.lower() in {".xlsx", ".xlsm"} and "app" in path.stem.lower()


def is_calendar_file(path):
    return (
        path.suffix.lower() in {".csv", ".xlsx", ".xlsm"}
        and not is_app_mapping_file(path)
        and bool(member_from_file(path))
    )


def load_rows(path):
    suffix = path.suffix.lower()
    if suffix == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            return list(csv.DictReader(handle))

    if suffix in {".xlsx", ".xlsm"}:
        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
        sheet = workbook.active
        header_row = next(sheet.iter_rows(min_row=1, max_row=1))
        headers = [
            str(cell.value).strip() if cell.value is not None else ""
            for cell in header_row
        ]
        rows = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not any(value is not None and str(value).strip() for value in row):
                continue
            rows.append(
                {
                    headers[index]: row[index] if index < len(row) else None
                    for index in range(len(headers))
                }
            )
        return rows

    return []


def load_app_mapping(source_dir):
    mapping_file = next(
        (path for path in sorted(source_dir.iterdir()) if is_app_mapping_file(path)),
        None,
    )
    if mapping_file is None:
        return {}, None

    rows = load_rows(mapping_file)
    mapping = {}
    for row in rows:
        code = normalize_tfs_code(row.get("App Code"))
        app_name = normalize_subject(row.get("App Name"))
        short_name = normalize_subject(row.get("Short Name"))
        if code and app_name:
            mapping[code] = {
                "appName": app_name,
                "shortName": short_name,
            }
    return mapping, {"file": mapping_file.name, "rows": len(rows), "mappedCodes": len(mapping)}


def subject_prefix(subject):
    text = re.sub(r"^Canceled:\s*", "", subject, flags=re.IGNORECASE).strip()
    match = re.match(r"^([A-Za-z]+)(?:\s*[-_:\[]|\s+|$)", text)
    return match.group(1).upper() if match else ""


def classify_subject(subject):
    text = normalize_subject(subject)
    lower = text.lower()
    prefix = subject_prefix(text)

    if lower.startswith("canceled:"):
        return "Canceled"
    if re.search(r"\bholiday\b", lower):
        return "Holiday"
    if re.search(r"\b(pto|annual leave|leave)\b", lower):
        return "PTO"
    if "no meeting required" in lower or "calendar reminder" in lower:
        return "Reminder"
    if prefix in {"PROJ", "PROJECT"}:
        return "Project"
    if prefix == "CR":
        return "CR"
    if prefix in {"MGMT", "MANAGEMENT"}:
        return "Mgmt"
    if prefix in {"SUP", "SUPPORT", "OPS", "INC"}:
        return "Sup"
    return "Other"


def split_work_name(value):
    text = normalize_subject(value)
    parts = re.split(r"\s+-\s+|[:|]|\uff1a", text, maxsplit=1)
    return parts[0].strip()


def extract_project_name(subject):
    text = re.sub(r"^Canceled:\s*", "", normalize_subject(subject), flags=re.IGNORECASE)
    bracket = re.search(r"\[([^\]]+)\]", text)
    if bracket:
        return bracket.group(1).strip()

    cleaned = re.sub(r"^(PROJ|PROJECT)\s*[-_:\s]+", "", text, flags=re.IGNORECASE)
    return split_work_name(cleaned) or "Unspecified Project"


def extract_cr_code(subject):
    return normalize_tfs_code(subject)


def extract_cr_fallback(subject):
    text = re.sub(r"^Canceled:\s*", "", normalize_subject(subject), flags=re.IGNORECASE)
    cleaned = re.sub(r"^CR\s*[-_:\s]+", "", text, flags=re.IGNORECASE)
    return split_work_name(cleaned) or "Unspecified CR"


def resolve_cr_system(subject, app_mapping):
    code = extract_cr_code(subject)
    if code and code in app_mapping:
        return app_mapping[code]["appName"], code, app_mapping[code]["appName"]
    if code:
        return code, code, ""
    fallback = extract_cr_fallback(subject)
    return fallback, "", ""


def work_item_name(subject, category, app_mapping):
    if category == "Project":
        return extract_project_name(subject)
    if category == "CR":
        cr_system, _code, _app_name = resolve_cr_system(subject, app_mapping)
        return cr_system
    return category


def show_time_label(value):
    text = str(value or "").strip()
    return SHOW_TIME_AS.get(text, text or "Unknown")


def month_end(day):
    if day.month == 12:
        return date(day.year + 1, 1, 1) - timedelta(days=1)
    return date(day.year, day.month + 1, 1) - timedelta(days=1)


def iter_days(start_day, end_day):
    current = start_day
    while current <= end_day:
        yield current
        current += timedelta(days=1)


def default_workdays_between(start_text, end_text):
    if not start_text or not end_text:
        return []
    start_day = date.fromisoformat(start_text)
    end_day = date.fromisoformat(end_text)
    days = []
    for current in iter_days(start_day, end_day):
        if current.weekday() < 5:
            days.append(current.isoformat())
    return days


def build_date_range(events):
    if not events:
        today = date.today()
        start = today.replace(day=1)
        return {
            "start": start.isoformat(),
            "end": month_end(start).isoformat(),
            "monthStart": start.isoformat(),
            "monthEnd": month_end(start).isoformat(),
            "eventStart": "",
            "eventEnd": "",
        }

    event_days = sorted(date.fromisoformat(event["date"]) for event in events)
    return {
        "start": event_days[0].isoformat(),
        "end": event_days[-1].isoformat(),
        "monthStart": event_days[0].replace(day=1).isoformat(),
        "monthEnd": month_end(event_days[-1]).isoformat(),
        "eventStart": event_days[0].isoformat(),
        "eventEnd": event_days[-1].isoformat(),
    }


def build_months(events):
    month_keys = sorted({event["date"][:7] for event in events})
    months = []
    for key in month_keys:
        year, month = [int(part) for part in key.split("-")]
        start = date(year, month, 1)
        months.append(
            {
                "value": key,
                "label": key,
                "start": start.isoformat(),
                "end": month_end(start).isoformat(),
            }
        )
    return months


def week_start(day):
    return day - timedelta(days=day.weekday())


def duration_hours(start, end):
    if end <= start:
        end += timedelta(days=1)
    return round((end - start).total_seconds() / 3600, 2), end


def is_tracked_event(event):
    return not event["canceled"] and not event["allDay"] and not event["isReminder"]


def build_events(app_mapping):
    events = []
    source_files = []
    event_id = 1

    for path in sorted(SOURCE_DIR.iterdir()):
        if not is_calendar_file(path):
            continue

        member = member_from_file(path)
        rows = load_rows(path)
        source_files.append({"member": member, "file": path.name, "rows": len(rows)})

        for row in rows:
            start_date = parse_date(row.get("Start Date"))
            end_date = parse_date(row.get("End Date"))
            if not start_date or not end_date:
                continue

            start = datetime.combine(start_date, parse_time(row.get("Start Time")))
            end = datetime.combine(end_date, parse_time(row.get("End Time")))
            hours, end = duration_hours(start, end)
            subject = normalize_subject(row.get("Subject"))
            category = classify_subject(subject)
            project_name = extract_project_name(subject) if category == "Project" else ""
            cr_system, cr_code, app_name = (
                resolve_cr_system(subject, app_mapping) if category == "CR" else ("", "", "")
            )
            all_day = bool_value(row.get("All day event"))
            canceled = subject.lower().startswith("canceled:")

            events.append(
                {
                    "id": event_id,
                    "member": member,
                    "sourceFile": path.name,
                    "date": start_date.isoformat(),
                    "weekStart": week_start(start_date).isoformat(),
                    "weekday": start.strftime("%a"),
                    "start": start.isoformat(timespec="minutes"),
                    "end": end.isoformat(timespec="minutes"),
                    "startTime": start.strftime("%H:%M"),
                    "endTime": end.strftime("%H:%M"),
                    "hours": hours,
                    "subject": subject,
                    "prefix": subject_prefix(subject),
                    "category": category,
                    "projectName": project_name,
                    "crCode": cr_code,
                    "appName": app_name,
                    "crSystem": cr_system,
                    "workItemName": work_item_name(subject, category, app_mapping),
                    "showTimeAs": show_time_label(row.get("Show time as")),
                    "organizer": normalize_subject(row.get("Meeting Organizer")),
                    "location": normalize_subject(row.get("Location")),
                    "allDay": all_day,
                    "canceled": canceled,
                    "isReminder": category == "Reminder",
                    "isWork": category in WORK_CATEGORIES,
                    "isTimeOff": category in TIME_OFF_CATEGORIES,
                    "isDistributionWork": category in DISTRIBUTION_CATEGORIES,
                }
            )
            event_id += 1

    return events, source_files


def category_hours(events):
    totals = Counter()
    for event in events:
        totals[event["category"]] += event["hours"]
    return {
        category: round(totals[category], 2)
        for category in CATEGORY_ORDER
        if totals[category]
    }


def build_week_targets(workdays):
    targets = Counter()
    for day_text in workdays:
        day = date.fromisoformat(day_text)
        targets[week_start(day).isoformat()] += STANDARD_DAY_HOURS
    return dict(sorted(targets.items()))


def build_weekly_summary(events, members, week_targets):
    weekly = []
    for member in members:
        for week, target_hours in week_targets.items():
            scoped = [
                event
                for event in events
                if event["member"] == member
                and event["weekStart"] == week
                and is_tracked_event(event)
            ]
            work_hours = round(
                sum(event["hours"] for event in scoped if event["isWork"]), 2
            )
            time_off_hours = round(
                sum(event["hours"] for event in scoped if event["isTimeOff"]), 2
            )
            credit_hours = round(work_hours + time_off_hours, 2)
            weekly.append(
                {
                    "member": member,
                    "weekStart": week,
                    "weekEnd": (
                        date.fromisoformat(week) + timedelta(days=6)
                    ).isoformat(),
                    "targetHours": round(target_hours, 2),
                    "workHours": work_hours,
                    "timeOffHours": time_off_hours,
                    "creditHours": credit_hours,
                    "varianceHours": round(credit_hours - target_hours, 2),
                    "categoryHours": category_hours(scoped),
                }
            )
    return weekly


def build_summary(events, workdays, members):
    summary = {}
    standard_hours = len(workdays) * STANDARD_DAY_HOURS

    for member in members:
        member_events = [event for event in events if event["member"] == member]
        tracked = [event for event in member_events if is_tracked_event(event)]
        work = [event for event in tracked if event["isWork"]]
        time_off = [event for event in tracked if event["isTimeOff"]]
        distribution = [event for event in tracked if event["isDistributionWork"]]
        work_hours = round(sum(event["hours"] for event in work), 2)
        distribution_hours = round(sum(event["hours"] for event in distribution), 2)
        time_off_hours = round(sum(event["hours"] for event in time_off), 2)
        credit_hours = round(work_hours + time_off_hours, 2)
        available_hours = round(standard_hours - time_off_hours, 2)

        summary[member] = {
            "rawEvents": len(member_events),
            "trackedEvents": len(tracked),
            "workEvents": len(work),
            "timeOffEvents": len(time_off),
            "standardHours": standard_hours,
            "workHours": work_hours,
            "distributionWorkHours": distribution_hours,
            "otherWorkHours": round(
                sum(event["hours"] for event in work if event["category"] == "Other"),
                2,
            ),
            "timeOffHours": time_off_hours,
            "availableWorkHours": available_hours,
            "creditHours": credit_hours,
            "varianceHours": round(credit_hours - standard_hours, 2),
            "utilization": round(work_hours / available_hours, 4)
            if available_hours
            else 0,
            "canceledEvents": sum(event["canceled"] for event in member_events),
            "allDayEvents": sum(event["allDay"] for event in member_events),
            "reminderEvents": sum(event["isReminder"] for event in member_events),
            "categoryHours": category_hours(tracked),
        }
    return summary


def main():
    app_mapping, app_mapping_source = load_app_mapping(SOURCE_DIR)
    events, source_files = build_events(app_mapping)
    members = sorted({event["member"] for event in events})
    months = build_months(events)
    date_range = build_date_range(events)
    workdays = default_workdays_between(date_range["start"], date_range["end"])
    week_targets = build_week_targets(workdays)
    default_month = months[0]["value"] if months else date_range["start"][:7]

    payload = {
        "generatedAt": datetime.now().isoformat(timespec="seconds"),
        "sourceDirectory": str(SOURCE_DIR),
        "sourceUrl": SOURCE_URL,
        "appMappingSource": app_mapping_source,
        "month": default_month,
        "monthLabel": default_month,
        "dateRange": date_range,
        "months": months,
        "workdayHours": STANDARD_DAY_HOURS,
        "standardWeekHours": STANDARD_WEEK_HOURS,
        "workdays": workdays,
        "weekTargets": week_targets,
        "categories": CATEGORY_ORDER,
        "workCategories": sorted(WORK_CATEGORIES),
        "timeOffCategories": sorted(TIME_OFF_CATEGORIES),
        "distributionCategories": DISTRIBUTION_CATEGORIES,
        "members": members,
        "sourceFiles": source_files,
        "assumptions": [
            "本报表按 Outlook 日历标题命名规范识别工时类别，主类仅包括 Project、CR、Sup、Mgmt，未匹配的工作事件归入 Other 供复核。",
            "Project 类按 Proj-[项目名] 识别，方括号中的内容作为项目 / 系统名称。",
            "CR 类按标题中的 TFS code 识别；若 App List 中存在映射，则报表显示 APP Name 替代 TFS code。",
            "Sup 类仅取标准前缀值，例如 SUP-INC、SUP-OPS、SUP-ADHOC、SUP-AI、SUP-RPA、SUP-Email；项目 / 系统栏仅保留空格前的 SUP-XXX。",
            "Mgmt 类仅取标准前缀值，例如 MGMT-REPORT、MGMT-TEAM、MGMT-PLAN；项目 / 系统栏仅保留空格前的 MGMT-XXX。",
            "时间记录按实际投入计算，支持晚间、周末和节假日工时；最小记录粒度参考 SOP 为 15 分钟。",
            "同一时间段在 Dashboard 中不处理时间重叠，按导入日历记录直接汇总。",
            "取消会议、全天事件和提醒类事件默认不计入工时汇总与 credit 口径。",
            "标准工时 = 当前所选范围内的配置工作日 × 8 小时；节假日从标准工时中扣除，补班日计入标准工时。",
            "工作负荷 = (工作时间 + PTO) / 标准工时；PTO 计入 credit，Holiday 不额外计入 credit。",
        ],
        "summary": build_summary(events, workdays, members),
        "weeklySummary": build_weekly_summary(events, members, week_targets),
        "events": events,
    }

    OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)
    content = "window.TIMESHEET_DATA = "
    content += json.dumps(payload, ensure_ascii=False, indent=2)
    content += ";\n"
    OUTPUT_FILE.write_text(content, encoding="utf-8")
    print(f"Wrote {OUTPUT_FILE}")
    print(f"Events: {len(events)}")
    if app_mapping_source:
        print(
            f"App mapping: {app_mapping_source['file']} "
            f"({app_mapping_source['mappedCodes']} codes)"
        )


if __name__ == "__main__":
    main()
